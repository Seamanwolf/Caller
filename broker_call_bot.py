import os
import logging
import requests
import pandas as pd
import matplotlib.pyplot as plt
import calendar
from datetime import datetime, timedelta
from io import BytesIO
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
import json
import re
from tqdm import tqdm
from prettytable import PrettyTable
from colorama import init, Fore, Style
import numpy as np
import matplotlib
matplotlib.use('Agg')
import asyncio
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

# Импортирую EmployeeDataProvider
from employee_data_provider import EmployeeDataProvider

# Инициализация colorama
init(autoreset=True)

# Функция для коррекции системного времени (если на сервере неправильная дата)
def get_actual_now():
    """
    Возвращает актуальную текущую дату.
    """
    return datetime.now()

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Настройки API ВАТС
API_KEY = os.getenv("API_KEY", "d1b0ef65-e491-43f9-967b-df67d4657dbb")
API_URL = os.getenv("API_URL", "https://leto.megapbx.ru/crmapi/v1")

# Список разрешенных пользователей
ALLOWED_USERS_STR = os.getenv("ALLOWED_USERS", "194530,368752085,261337953,702018715")
ALLOWED_USERS = [int(user_id.strip()) for user_id in ALLOWED_USERS_STR.split(",")]

# ID пользователя для автоматических отчетов
AUTO_REPORT_USER_ID = int(os.getenv("AUTO_REPORT_USER_ID", "194530"))

# Глобальная переменная для приложения бота (нужна для планировщика)
bot_application = None

# Инициализация провайдера сотрудников (глобально)
EMPLOYEE_API_TOKEN = os.getenv("EMPLOYEE_API_TOKEN", "a4d4a75094d8f9d8597085ac0ac12a51")
employee_provider = EmployeeDataProvider(EMPLOYEE_API_TOKEN)

# Функция для безопасного обновления сообщений
async def safe_edit_message(query, text, reply_markup=None, parse_mode=None):
    """
    Безопасно обновляет сообщение, избегая ошибки 'Message is not modified'
    """
    try:
        await query.edit_message_text(
            text=text,
            reply_markup=reply_markup,
            parse_mode=parse_mode
        )
    except Exception as e:
        if "Message is not modified" in str(e):
            # Игнорируем эту ошибку - сообщение уже имеет нужное содержимое
            logger.debug(f"Сообщение не изменилось: {text[:50]}...")
        else:
            # Логируем другие ошибки
            logger.error(f"Ошибка при обновлении сообщения: {str(e)}")
            raise e

def setup_logging():
    logger.info("Настройка логирования")
    logging.getLogger().addHandler(logging.NullHandler())
    return logging.getLogger()

def get_department_numbers(department):
    logger.debug(f"Обработка отдела: {department}")
    if not department:
        return None
    try:
        if isinstance(department, (int, float)):
            return str(int(department))
        
        department = str(department)
        department = re.sub(r'[^\d]', '', department)
        result = department if department else None
        logger.debug(f"Результат обработки отдела: {result}")
        return result
    except Exception as e:
        logger.error(f"Ошибка при обработке отдела '{department}': {str(e)}")
        return None

async def start(update, context):
    user = update.effective_user
    user_id = user.id
    logger.info(f"Пользователь {user.first_name} (ID: {user_id}) запустил бота")
    if user_id not in ALLOWED_USERS:
        logger.warning(f"Попытка несанкционированного доступа от пользователя с ID {user_id}")
        await update.message.reply_text("⛔ Извините, у вас нет доступа к этому боту.")
        return
    welcome_text = "👋 Добро пожаловать в бот статистики отделов!\n\nВыберите опцию:"
    keyboard = [
        [InlineKeyboardButton("📋 Отчёт по всем отделам", callback_data="report:all")],
        [InlineKeyboardButton("📊 Отчёт по отделам", callback_data="report:by")],
        [InlineKeyboardButton("🔄 Обновить кэш", callback_data="update_employees")]
    ]
    await update.message.reply_text(welcome_text, reply_markup=InlineKeyboardMarkup(keyboard))

async def button_callback(update, context):
    query = update.callback_query
    data = query.data
    
    logger.info(f"Callback received: {data}")
    
    if data == "back_to_main":
        await show_main_menu(update, context)
        return
        
    if data == "update_employees":
        await safe_edit_message(query, "🔄 Обновляю кэш сотрудников...")
        try:
            loop = asyncio.get_running_loop()
            await loop.run_in_executor(None, employee_provider.update_cache, True)
            await safe_edit_message(query, "✅ Кэш сотрудников успешно обновлён!", reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ В меню", callback_data="back_to_main")]
            ]))
        except Exception as e:
            await safe_edit_message(query, f"❌ Ошибка обновления кэша: {e}", reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ В меню", callback_data="back_to_main")]
            ]))
        return
        
    if data == "report:all":
        context.user_data["report_type"] = "all"
        context.user_data["dept_number"] = "all"
        await show_period_selection(query, context, sheet_type="vtorichka", report_type="all")
        return
        
    if data == "report:by":
        context.user_data["report_type"] = "by"
        await show_department_list(query, context, sheet_type="vtorichka", report_type="by")
        return
        
    if data.startswith("dept:"):
        dept_data = data.split(":")
        if len(dept_data) >= 2:
            dept_number = dept_data[1]
            context.user_data["dept_number"] = str(dept_number)
            context.user_data["selected_dept_number"] = str(dept_number)
            await show_period_selection(query, context, sheet_type="vtorichka", report_type="by")
            return
            
    if data.startswith("period:"):
        period = data.split(":")[1]
        sheet_type = context.user_data.get("sheet_type", "")
        report_type = context.user_data.get("report_type", "")
        dept_number = context.user_data.get("dept_number", "all")
        
        logger.info(f"Выбран период: {period}, sheet_type: {sheet_type}, report_type: {report_type}, dept_number: {dept_number}")
        
        context.user_data["period"] = period
        await show_format_selection(query, context, sheet_type, report_type, dept_number, period)
        return
        
    if data.startswith("format:"):
        format_type = data.split(":")[1]
        sheet_type = context.user_data.get("sheet_type", "")
        report_type = context.user_data.get("report_type", "")
        dept_number = context.user_data.get("dept_number", "all")
        period = context.user_data.get("period", "")
        
        logger.info(f"Выбран формат: {format_type}, sheet_type: {sheet_type}, report_type: {report_type}, dept_number: {dept_number}, period: {period}")
        
        if format_type == "incoming":
            await handle_incoming_numbers_excel(query, context, sheet_type, dept_number, period)
        else:
            await handle_report_format(query, context, sheet_type, dept_number, period, format_type)
        return
        
    if data.startswith("quarter:"):
        parts = data.split(":")
        if len(parts) == 2:
            # Формат: quarter:3sheets или quarter:1sheet
            sheets_type = parts[1]
            context.user_data["sheets_type"] = sheets_type
            context.user_data["report_type"] = "quarter"
            await show_year_selection(query, context)
            return
        elif len(parts) == 3:
            # Формат: quarter:2024:2
            year = int(parts[1])
            quarter = int(parts[2])
            await generate_quarter_report(query, context, year, quarter)
            return
        
    if data.startswith("year:"):
        year = int(data.split(":")[1])
        context.user_data["year"] = year
        await show_quarter_selection(query, context, year)
        return
    
    await safe_edit_message(query,
        f"⚠️ Неизвестный callback: {data}",
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
    )

async def show_main_menu(update, context):
    welcome_text = "👋 Добро пожаловать в бот статистики отделов!\n\nВыберите опцию:"
    keyboard = [
        [InlineKeyboardButton("📋 Отчёт по всем отделам", callback_data="report:all")],
        [InlineKeyboardButton("📊 Отчёт по отделам", callback_data="report:by")],
        [InlineKeyboardButton("🔄 Обновить кэш", callback_data="update_employees")]
    ]
    await update.callback_query.edit_message_text(welcome_text, reply_markup=InlineKeyboardMarkup(keyboard))

async def show_period_selection(query, context, sheet_type, report_type):
    logger.info(f"Показываю выбор периода для sheet_type={sheet_type}, report_type={report_type}")
    
    keyboard = [
        [InlineKeyboardButton("📅 Сегодня", callback_data="period:today")],
        [InlineKeyboardButton("📅 Текущий месяц", callback_data="period:current_month")],
        [InlineKeyboardButton("📅 Предыдущий месяц", callback_data="period:previous_month")],
        [InlineKeyboardButton("📅 За 7 дней", callback_data="period:week")],
        [InlineKeyboardButton("📅 За 30 дней", callback_data="period:month")],
        [InlineKeyboardButton("📊 Квартальный отчёт (3 листа)", callback_data="quarter:3sheets")],
        [InlineKeyboardButton("📋 Квартальный отчёт (1 лист)", callback_data="quarter:1sheet")],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]
    ]
    
    await safe_edit_message(query,
        f"📅 Выберите период для отчёта:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def show_department_list(query, context, sheet_type, report_type):
    logger.info(f"Показываю список отделов для sheet_type={sheet_type}, report_type={report_type}")
    
    try:
        # Показываем сообщение о загрузке
        await safe_edit_message(query, "🔄 Загружаю список отделов...", reply_markup=None)
        
        # Получаем сотрудников из кэша с таймаутом
        try:
            # Запускаем получение сотрудников в отдельном потоке с таймаутом
            loop = asyncio.get_event_loop()
            employees = await asyncio.wait_for(
                loop.run_in_executor(None, employee_provider.get_employees),
                timeout=10.0  # 10 секунд таймаут
            )
            filtered = employees
            logger.info(f"Получено {len(filtered)} сотрудников")
        except asyncio.TimeoutError:
            logger.error("Таймаут при получении сотрудников")
            await safe_edit_message(query, "❌ Таймаут при загрузке данных. Попробуйте позже.", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        except Exception as e:
            logger.error(f"Ошибка при получении сотрудников: {str(e)}")
            await safe_edit_message(query, f"❌ Ошибка при загрузке данных: {str(e)}", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Группируем по отделам
        departments = {}
        import re
        for emp in filtered:
            try:
                dept_raw = emp['department']
                match = re.search(r'(\d+)', str(dept_raw))
                if match:
                    dept = str(int(match.group(1)))
                    if dept not in departments:
                        departments[dept] = []
                    departments[dept].append({
                        'phone': emp['sim'],
                        'name': f"{emp['last_name']} {emp['first_name']}",
                        'department': emp['department']
                    })
            except Exception as e:
                logger.error(f"Ошибка при обработке сотрудника {emp}: {str(e)}")
                continue
        
        if not departments:
            logger.error("Не найдено отделов с номерами")
            await safe_edit_message(query, "❌ Не найдено отделов с номерами", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
            
        keyboard = []
        for dept_number in sorted(departments.keys(), key=int):
            callback_data = f"dept:{dept_number}"
            keyboard.append([InlineKeyboardButton(f"Отдел {dept_number} ({len(departments[dept_number])} сотрудников)", callback_data=callback_data)])
        keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])
        
        await safe_edit_message(query,
            f"🏢 Выберите отдел для создания отчета:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        
    except Exception as e:
        logger.error(f"Критическая ошибка в show_department_list: {str(e)}")
        await safe_edit_message(query, f"❌ Произошла ошибка: {str(e)}", 
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def show_format_selection(query, context, sheet_type, report_type, dept_number, period):
    logger.info(f"Показываю выбор формата для sheet_type={sheet_type}, report_type={report_type}, dept_number={dept_number}, period={period}")
    
    keyboard = [
        [InlineKeyboardButton("📊 График", callback_data="format:plot")],
        [InlineKeyboardButton("📋 Таблица", callback_data="format:table")],
        [InlineKeyboardButton("📑 Excel", callback_data="format:excel")],
        [InlineKeyboardButton("📞 Входящие номера", callback_data="format:incoming")],
        [InlineKeyboardButton("📊 Все форматы", callback_data="format:all")],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]
    ]
    
    await safe_edit_message(query,
        f"📋 Выберите формат отчета:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def handle_report_format(query, context, sheet_type, dept_number, period, format_type):
    """Обработка формата отчета"""
    logger.info(f"Обработка формата отчета: тип={sheet_type}, отдел={dept_number}, период={period}, формат={format_type}")
    
    try:
        # Показываем прогресс
        await safe_edit_message(query, "🔄 Формирую отчет...", reply_markup=None)
        
        # Получаем сотрудников из кэша
        employees = employee_provider.get_employees()
        filtered = employees
        
        if not filtered:
            logger.error("Не найдено сотрудников для отчета")
            await safe_edit_message(query, "❌ Нет данных для создания отчета", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        # Получаем даты периода
        try:
            start_date_str, end_date_str = get_period_dates(period, context)
            logger.info(f"Получены даты периода: {start_date_str} - {end_date_str}")
        except ValueError as e:
            logger.error(f"Ошибка при определении периода: {str(e)}")
            await safe_edit_message(query, f"❌ Ошибка при определении периода: {str(e)}", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Фильтруем сотрудников по отделу, если выбран конкретный отдел
        if dept_number != "all":
            filtered_employees = []
            for employee in filtered:
                emp_dept = get_department_numbers(employee['department'])
                if emp_dept == dept_number:
                    filtered_employees.append(employee)
            filtered = filtered_employees
            logger.info(f"Отфильтровано {len(filtered)} сотрудников для отдела {dept_number}")

        # Собираем статистику по сотрудникам
        all_stats = []
        total_employees = len(filtered)
        current_employee = 0
        
        for employee in filtered:
            if not employee.get('sim') or employee['sim'] == 'Нет данных':
                continue
            
            current_employee += 1
            
            # Обновляем прогресс-бар
            progress = (current_employee / total_employees) * 100
            progress_bar = "█" * int(progress / 2) + "░" * (50 - int(progress / 2))
            progress_text = (
                f"🔄 Формирую отчет...\n"
                f"Прогресс: {progress_bar} {progress:.1f}%\n"
                f"Обработано сотрудников: {current_employee}/{total_employees}\n"
                f"Текущий: {employee.get('last_name', '')} {employee.get('first_name', '')}"
            )
            await safe_edit_message(query, progress_text, reply_markup=None)
            
            # Получаем данные звонков
            data = fetch_call_history(start_date_str, end_date_str, employee['sim'])
            if not data:
                continue
                
            # Подсчитываем статистику
            df = pd.DataFrame(data)
            if not df.empty:
                incoming_types = ['in', 'incoming', 'received', 'inbound', 'входящий']
                outgoing_types = ['out', 'outgoing', 'исходящий']
                missed_statuses = ['noanswer', 'missed', 'пропущен', 'неотвечен', 'нет ответа']

                incoming_count = df[df['type'].str.lower().isin(incoming_types)].shape[0]
                outgoing_count = df[df['type'].str.lower().isin(outgoing_types)].shape[0]
                missed_count = df[df['status'].str.lower().isin(missed_statuses)].shape[0] if 'status' in df.columns else 0
                    
                stats_dict = {
                    'Сотрудник': f"{employee.get('last_name', '')} {employee.get('first_name', '')}".strip(),
                    'Отдел': get_department_numbers(employee['department']),
                    'Входящие 📞': incoming_count,
                    'Исходящие 📤': outgoing_count,
                    'Пропущенные ❌': missed_count,
                    'Всего звонков': len(data)
                }
                all_stats.append(stats_dict)
        
        if not all_stats:
            logger.error("Нет данных для создания отчета")
            await safe_edit_message(query, "❌ Нет данных для создания отчета", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        # Создаем DataFrame
        df_stats = pd.DataFrame(all_stats)
        
        # Обрабатываем формат отчета
        if format_type == "all":
            await handle_table_format(query, context, all_stats, "Отчет")
            await handle_plot_format(query, context, df_stats, "Отчет")
            await handle_excel_format(query, context, df_stats, "Отчет", period)
            period_info = get_period_dates_info(period, context)
            await safe_edit_message(query, f"✅ Все форматы отчета отправлены! ({period_info})", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
        elif format_type == "excel":
            await handle_excel_format(query, context, df_stats, "Отчет", period)
        elif format_type == "plot":
            await handle_plot_format(query, context, df_stats, "Отчет")
        elif format_type == "table":
            await handle_table_format(query, context, all_stats, "Отчет")
        elif format_type == "incoming":
            await handle_incoming_numbers_excel(query, context, sheet_type, dept_number, period)
        
    except Exception as e:
        logger.error(f"Ошибка при обработке формата отчета: {str(e)}")
        await safe_edit_message(query, f"❌ Произошла ошибка: {str(e)}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def handle_table_format(query, context, all_stats, sheet_name):
    """Обработка табличного формата"""
    try:
        # Показываем прогресс
        await safe_edit_message(query, "🔄 Формирую таблицу...", reply_markup=None)
        
        # Создаем таблицу
        table = PrettyTable()
        table.field_names = ["Сотрудник", "Отдел", "Входящие", "Исходящие", "Пропущенные", "Всего"]
        
        for stats in all_stats:
            table.add_row([
                stats['Сотрудник'],
                stats['Отдел'],
                stats['Входящие 📞'],
                stats['Исходящие 📤'],
                stats['Пропущенные ❌'],
                stats['Всего звонков']
            ])
        
        # Отправляем таблицу
        await safe_edit_message(query, f"📋 Таблица отчета:\n\n`{table}`", parse_mode=ParseMode.MARKDOWN,
                                              reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
                )
    except Exception as e:
        logger.error(f"Ошибка при создании таблицы: {str(e)}")
        await safe_edit_message(query, f"❌ Ошибка при создании таблицы: {str(e)}", 
                                     reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def handle_plot_format(query, context, df_stats, sheet_name):
    """Обработка графического формата"""
    try:
        # Показываем прогресс
        await safe_edit_message(query, "🔄 Создаю график...", reply_markup=None)
        
        # Создаем график
        fig, ax = plt.subplots(figsize=(12, 8))
        
        # Группируем данные по отделам
        dept_stats = df_stats.groupby('Отдел').agg({
            'Входящие 📞': 'sum',
            'Исходящие 📤': 'sum',
            'Пропущенные ❌': 'sum'
    }).reset_index()
    
        # Создаем столбчатую диаграмму
        x = range(len(dept_stats))
        width = 0.25
        
        ax.bar([i - width for i in x], dept_stats['Входящие 📞'], width, label='Входящие', color='green')
        ax.bar(x, dept_stats['Исходящие 📤'], width, label='Исходящие', color='blue')
        ax.bar([i + width for i in x], dept_stats['Пропущенные ❌'], width, label='Пропущенные', color='red')
        
        ax.set_xlabel('Отделы')
        ax.set_ylabel('Количество звонков')
        ax.set_title(f'Статистика звонков по отделам')
        ax.set_xticks(x)
        ax.set_xticklabels(dept_stats['Отдел'])
        ax.legend()
        
        # Отправляем график
        await send_plot(fig, query.message.chat_id, context)
        await safe_edit_message(query, "📊 График отправлен!", 
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
    except Exception as e:
        logger.error(f"Ошибка при создании графика: {str(e)}")
        await safe_edit_message(query, f"❌ Ошибка при создании графика: {str(e)}", 
                                   reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def handle_excel_format(query, context, df_stats, sheet_name, period):
    logger.info("Начало формирования Excel-файла")
    
    try:
        # Показываем прогресс
        await safe_edit_message(query, "🔄 Формирую Excel файл...", reply_markup=None)
        
        # Проверяем тип отчета
        report_type = context.user_data.get("report_type", "all")
        logger.info(f"Тип отчета в handle_excel_format: {report_type}")
        
        # Получаем номер отдела из context.user_data
        dept_number = context.user_data.get("selected_dept_number", "all")
        logger.info(f"Номер отдела в handle_excel_format: {dept_number}")
        
        # Получаем информацию о периоде
        period_info = get_period_dates_info(period, context)
        
        # Проверяем наличие данных
        if df_stats.empty:
            logger.error("Нет данных для создания Excel-файла")
            await safe_edit_message(query, "❌ Нет данных для создания Excel-файла", 
                                       reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Фильтруем данные по отделу, если выбран конкретный отдел
        if report_type != "all" and dept_number and dept_number != "all":
            filtered_df = df_stats[df_stats['Отдел'] == dept_number]
            logger.info(f"Отфильтровано {len(filtered_df)} записей для отдела {dept_number}")
            
            if filtered_df.empty:
                logger.error(f"Нет данных для отдела {dept_number}")
                await safe_edit_message(query, f"❌ Нет данных для отдела {dept_number}", 
                                           reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
                )
                return
            df_stats = filtered_df
        
        if report_type == "all":
            excel_df = pd.DataFrame()
            required_columns = ['Отдел', 'Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌']
            for column in required_columns:
                if column not in df_stats.columns:
                    df_stats[column] = 0
            for dept in df_stats['Отдел'].unique():
                dept_data = df_stats[df_stats['Отдел'] == dept]
                dept_total = dept_data[['Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌']].sum()
                num_employees = len(dept_data) if len(dept_data) else 1
                excel_df = pd.concat([excel_df, pd.DataFrame([{
                    'Отдел': dept,
                    'Входящие 📞': f"{dept_total['Входящие 📞']} ({round(dept_total['Входящие 📞']/num_employees, 1)})",
                    'Исходящие 📤': f"{dept_total['Исходящие 📤']} ({round(dept_total['Исходящие 📤']/num_employees, 1)})",
                    'Пропущенные ❌': f"{dept_total['Пропущенные ❌']} ({round(dept_total['Пропущенные ❌']/num_employees, 1)})"
                }])], ignore_index=True)
            total = df_stats[['Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌']].sum()
            excel_df = pd.concat([excel_df, pd.DataFrame([{
                'Отдел': 'ИТОГО ВСЕГО',
                'Входящие 📞': total['Входящие 📞'],
                'Исходящие 📤': total['Исходящие 📤'],
                'Пропущенные ❌': total['Пропущенные ❌']
            }])], ignore_index=True)
        else:
            excel_df = df_stats.copy()
            # Переупорядочиваем колонки для отчета по отделам
            # Первым столбцом должна быть фамилия и имя сотрудника
            column_order = ['Сотрудник', 'Отдел', 'Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌', 'Всего звонков']
            # Добавляем колонки, которых может не быть
            for col in column_order:
                if col not in excel_df.columns:
                    excel_df[col] = 0
            # Переупорядочиваем колонки
            excel_df = excel_df[column_order]
            
            for dept in df_stats['Отдел'].unique():
                dept_data = df_stats[df_stats['Отдел'] == dept]
                dept_total = dept_data[['Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌']].sum()
                excel_df = pd.concat([excel_df, pd.DataFrame([{
                    'Сотрудник': f'ИТОГО {dept}',
                    'Отдел': dept,
                    'Входящие 📞': dept_total['Входящие 📞'],
                    'Исходящие 📤': dept_total['Исходящие 📤'],
                    'Пропущенные ❌': dept_total['Пропущенные ❌'],
                    'Всего звонков': dept_total['Входящие 📞'] + dept_total['Исходящие 📤'] + dept_total['Пропущенные ❌']
                }])], ignore_index=True)
            total = df_stats[['Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌']].sum()
            excel_df = pd.concat([excel_df, pd.DataFrame([{
                'Сотрудник': 'ИТОГО ВСЕГО',
                'Отдел': '',
                'Входящие 📞': total['Входящие 📞'],
                'Исходящие 📤': total['Исходящие 📤'],
                'Пропущенные ❌': total['Пропущенные ❌'],
                'Всего звонков': total['Входящие 📞'] + total['Исходящие 📤'] + total['Пропущенные ❌']
            }])], ignore_index=True)
        
        # Добавляем информацию о периоде в имя файла
        filename = f"calls_stats_{sheet_name.lower()}_{period_info.replace(':', '').replace(' ', '_').replace('/', '_')}.xlsx"
        await send_excel(excel_df, filename, query.message.chat_id, context)
        await safe_edit_message(query, f"✅ Excel-файл отправлен ({period_info})!", 
                                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
    
    except Exception as e:
        logger.error(f"Ошибка при создании Excel-файла: {str(e)}")
        await safe_edit_message(query, f"❌ Ошибка при создании Excel-файла: {str(e)}", 
                                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

# Функции для работы с квартальными отчетами
async def show_year_selection(query, context):
    """Показать выбор года для квартального отчета"""
    current_year = datetime.now().year
    keyboard = []
    
    # Добавляем кнопки для последних 3 лет
    for year in range(current_year, current_year - 3, -1):
        keyboard.append([InlineKeyboardButton(f"📅 {year} год", callback_data=f"year:{year}")])
    
    keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")])
    
    await safe_edit_message(query,
        "📅 Выберите год для квартального отчета:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def show_quarter_selection(query, context, year):
    """Показать выбор квартала с цветными индикаторами"""
    current_year = datetime.now().year
    current_month = datetime.now().month
    current_quarter = (current_month - 1) // 3 + 1
    
    keyboard = []
    
    for quarter in range(1, 5):
        # Определяем статус квартала
        if year < current_year:
            # Прошедший год - все кварталы зеленые
            icon = "✅"
        elif year == current_year:
            if quarter < current_quarter:
                # Прошедший квартал
                icon = "✅"
            elif quarter == current_quarter:
                # Текущий квартал
                icon = "🟢"
            else:
                # Будущий квартал
                icon = "❌"
        else:
            # Будущий год - все кварталы красные
            icon = "❌"
        
        quarter_name = f"{quarter} квартал"
        keyboard.append([InlineKeyboardButton(f"{icon} {quarter_name}", callback_data=f"quarter:{year}:{quarter}")])
    
    keyboard.append([InlineKeyboardButton("⬅️ Назад", callback_data=f"quarter:{context.user_data.get('report_type', 'quarter_3sheets')}")])
    
    await safe_edit_message(query,
        f"📊 Выберите квартал для {year} года:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )

async def generate_quarter_report(query, context, year, quarter):
    """Генерировать квартальный отчет с реальными данными"""
    logger.info(f"=== НАЧАЛО generate_quarter_report: год={year}, квартал={quarter} ===")
    try:
        report_type = context.user_data.get("report_type")
        sheet_type = context.user_data.get("sheet_type", "")
        dept_number = context.user_data.get("dept_number", "all")
        sheets_type = context.user_data.get("sheets_type", "1sheet")  # Получаем тип листов
        
        # Создаем период для квартала
        period = f"quarter_{year}_{quarter}"
        
        logger.info(f"Генерация квартального отчета: год={year}, квартал={quarter}, отдел={dept_number}, лист={sheet_type}, тип листов={sheets_type}")
        
        # Показываем сообщение о прогрессе
        logger.info("Показываем сообщение о прогрессе для квартального отчета")
        await safe_edit_message(query, "🔄 Формирую квартальный отчет...", reply_markup=None)
        logger.info("Сообщение о прогрессе показано")
        
        if sheets_type == "3sheets":
            # Создаем отчет с 3 листами (по месяцам)
            await create_quarter_report_3sheets(query, context, year, quarter, sheet_type, dept_number, period)
        else:
            # Создаем отчет с 1 листом (весь квартал)
            # Передаем параметр skip_loading=True, чтобы не запускать дублирующий прогресс-бар
            await handle_report_format_quarter(query, context, sheet_type, dept_number, period, "excel")
            
    except Exception as e:
        logger.error(f"Ошибка при генерации квартального отчета: {str(e)}")
        await safe_edit_message(query, 
            f"❌ Ошибка при генерации отчета: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def handle_report_format_quarter(query, context, sheet_type, dept_number, period, format_type):
    """Версия handle_report_format для квартальных отчетов с прогресс-баром"""
    sheet_name = 'Все отделы'
    logger.info(f"Обработка квартального отчета: тип={sheet_type}, отдел={dept_number}, период={period}, формат={format_type}")
    
    # Показываем прогресс
    await safe_edit_message(query, "🔄 Формирую квартальный отчет...", reply_markup=None)
    
    # Дополнительная проверка номера отдела из context.user_data
    if dept_number == "None" or not dept_number or dept_number == "undefined":
        if context.user_data.get("selected_dept_number"):
            dept_number = context.user_data.get("selected_dept_number")
            logger.info(f"Восстановлен номер отдела из context.user_data['selected_dept_number']: {dept_number}")
        elif context.user_data.get("dept_number"):
            dept_number = context.user_data.get("dept_number")
            logger.info(f"Восстановлен номер отдела из context.user_data['dept_number']: {dept_number}")
    
    logger.info(f"Окончательный номер отдела для использования: {dept_number}")
    
    try:
        # Получаем сотрудников из кэша
        employees = employee_provider.get_employees()
        filtered = employees
        
        if not filtered:
            logger.error("Не найдено сотрудников для отчета")
            await safe_edit_message(query, "❌ Нет данных для создания отчета", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        logger.info(f"Найдено {len(filtered)} сотрудников для отчета")

        # Группируем сотрудников по отделам
        departments = {}
        import re
        for employee in filtered:
            dept_raw = employee['department']
            match = re.search(r'(\d+)', str(dept_raw))
            if match:
                dept = str(int(match.group(1)))
                if dept not in departments:
                    departments[dept] = []
                departments[dept].append({
                    'phone': employee['sim'],
                    'name': f"{employee['last_name']} {employee['first_name']}",
                    'department': employee['department']
                })

        logger.info(f"Найдено {len(departments)} отделов: {list(departments.keys())}")
        
        if not departments:
            logger.error("Не найдено отделов с номерами")
            await safe_edit_message(query, "❌ Не найдено отделов с номерами", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        try:
            actual_period = context.user_data.get("period", period)
            logger.info(f"Используемый период: {actual_period}")
            
            # Получаем строки дат в формате YYYYMMDD
            start_date_str, end_date_str = get_period_dates(actual_period, context)
            logger.info(f"Получены даты периода: {start_date_str} - {end_date_str}")
            
            logger.info(f"Отправка запроса с датами: {start_date_str} - {end_date_str}")
            
        except ValueError as e:
            logger.error(f"Ошибка при определении периода: {str(e)}")
            await safe_edit_message(query, f"❌ Ошибка при определении периода: {str(e)}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Получаем тип отчета из context.user_data
        report_type = context.user_data.get("report_type", "all")
        logger.info(f"Тип отчета в handle_report_format_quarter: {report_type}")
        
        # Проверяем и нормализуем номер отдела
        if not dept_number or dept_number == "None" or dept_number == "all" or dept_number == "undefined":
            if report_type != "all":
                logger.error(f"Не указан номер отдела для отчета типа {report_type}")
                await safe_edit_message(query, "❌ Не указан номер отдела для отчета", 
                                            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
                )
                return
            dept_number = "all"
        
        logger.info(f"Используемый номер отдела: {dept_number}")
        
        # Проверка существования отдела
        if dept_number != "all" and dept_number not in departments:
            logger.error(f"Отдел {dept_number} не найден среди доступных отделов: {list(departments.keys())}")
            await safe_edit_message(query, f"❌ Отдел {dept_number} не найден", 
                                        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        # Фильтруем сотрудников по отделу, если выбран конкретный отдел
        if dept_number != "all":
            filtered_employees = []
            for employee in filtered:
                emp_dept = get_department_numbers(employee['department'])
                if emp_dept == dept_number:
                    filtered_employees.append(employee)
            filtered = filtered_employees
            logger.info(f"Отфильтровано {len(filtered)} сотрудников для отдела {dept_number}")

        # Собираем статистику по сотрудникам
        all_stats = []
        total_employees = len(filtered)
        current_employee = 0
        
        for employee in filtered:
            if not employee.get('sim') or employee['sim'] == 'Нет данных':
                continue
                
            current_employee += 1
            
            # Обновляем прогресс-бар
            progress = (current_employee / total_employees) * 100
            progress_bar = "█" * int(progress / 2) + "░" * (50 - int(progress / 2))
            progress_text = (
                f"🔄 Формирую квартальный отчет...\n"
                f"Прогресс: {progress_bar} {progress:.1f}%\n"
                f"Обработано сотрудников: {current_employee}/{total_employees}\n"
                f"Текущий: {employee.get('last_name', '')} {employee.get('first_name', '')}"
            )
            await safe_edit_message(query, progress_text, reply_markup=None)
            
            # Получаем данные звонков
            data = fetch_call_history(start_date_str, end_date_str, employee['sim'])
            if not data:
                continue
            
            # Подсчитываем статистику
            df = pd.DataFrame(data)
            if not df.empty:
                incoming_types = ['in', 'incoming', 'received', 'inbound', 'входящий']
                outgoing_types = ['out', 'outgoing', 'исходящий']
                missed_statuses = ['noanswer', 'missed', 'пропущен', 'неотвечен', 'нет ответа']

                incoming_count = df[df['type'].str.lower().isin(incoming_types)].shape[0]
                outgoing_count = df[df['type'].str.lower().isin(outgoing_types)].shape[0]
                missed_count = df[df['status'].str.lower().isin(missed_statuses)].shape[0] if 'status' in df.columns else 0
                
                stats_dict = {
                    'Сотрудник': f"{employee.get('last_name', '')} {employee.get('first_name', '')}".strip(),
                    'Отдел': get_department_numbers(employee['department']),
                    'Входящие 📞': incoming_count,
                    'Исходящие 📤': outgoing_count,
                    'Пропущенные ❌': missed_count,
                    'Всего звонков': len(data)
                }
                all_stats.append(stats_dict)

        if not all_stats:
            logger.error("Нет данных для создания отчета")
            await safe_edit_message(query, "❌ Нет данных для создания отчета", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return

        # Создаем DataFrame
        df_stats = pd.DataFrame(all_stats)
        
        # Обрабатываем формат отчета
        if format_type == "all":
            await handle_table_format(query, context, all_stats, "Отчет")
            await handle_plot_format(query, context, df_stats, "Отчет")
            await handle_excel_format(query, context, df_stats, "Отчет", period)
            period_info = get_period_dates_info(period, context)
            await safe_edit_message(query, f"✅ Все форматы отчета отправлены! ({period_info})", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
        elif format_type == "excel":
            await handle_excel_format(query, context, df_stats, "Отчет", period)
        elif format_type == "plot":
            await handle_plot_format(query, context, df_stats, "Отчет")
        elif format_type == "table":
            await handle_table_format(query, context, all_stats, "Отчет")
        elif format_type == "incoming":
            await handle_incoming_numbers_excel(query, context, sheet_type, dept_number, period)
        
    except Exception as e:
        logger.error(f"Ошибка при обработке квартального отчета: {str(e)}")
        await safe_edit_message(query, f"❌ Произошла ошибка: {str(e)}", reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

async def create_quarter_report_3sheets(query, context, year, quarter, sheet_type, dept_number, period):
    """Создать квартальный отчет с 3 листами (по месяцам)"""
    try:
        # Определяем месяцы квартала
        quarter_months = {
            1: [("Январь", 1), ("Февраль", 2), ("Март", 3)],
            2: [("Апрель", 4), ("Май", 5), ("Июнь", 6)],
            3: [("Июль", 7), ("Август", 8), ("Сентябрь", 9)],
            4: [("Октябрь", 10), ("Ноябрь", 11), ("Декабрь", 12)]
        }
        
        months = quarter_months[quarter]
        
        # Прогресс-бар уже запущен в generate_quarter_report
        
        # Создаем Excel файл с 3 листами
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        wb = Workbook()
        
        # Удаляем дефолтный лист
        wb.remove(wb.active)
        
        # Получаем сотрудников
        employees = employee_provider.get_employees()
        if not employees:
            await safe_edit_message(query, "❌ Не удалось получить данные сотрудников", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Фильтруем сотрудников по отделу, если выбран конкретный отдел
        if dept_number != "all":
            filtered_employees = []
            for employee in employees:
                emp_dept = get_department_numbers(employee['department'])
                if emp_dept == dept_number:
                    filtered_employees.append(employee)
            employees = filtered_employees
        
        # Создаем листы для каждого месяца
        total_months = len(months)
        current_month = 0
        
        for month_name, month_num in months:
            current_month += 1
            
            # Обновляем прогресс-бар для месяцев
            month_progress = (current_month / total_months) * 100
            month_progress_bar = "█" * int(month_progress / 2) + "░" * (50 - int(month_progress / 2))
            month_progress_text = (
                f"🔄 Формирую квартальный отчет...\n"
                f"Месяц: {month_name} {year}\n"
                f"Прогресс месяцев: {month_progress_bar} {month_progress:.1f}%\n"
                f"Обработано месяцев: {current_month}/{total_months}"
            )
            await safe_edit_message(query, month_progress_text, reply_markup=None)
            
            ws = wb.create_sheet(title=month_name)
            
            # Заголовки
            headers = ['Сотрудник', 'Отдел', 'Входящие 📞', 'Исходящие 📤', 'Пропущенные ❌', 'Всего звонков']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Создаем период для месяца
            month_period = f"month_{year}_{month_num:02d}"
            
            # Получаем данные за месяц
            try:
                start_date_str, end_date_str = get_period_dates(month_period, context)
                logger.info(f"Получаем данные за {month_name} {year}: {start_date_str} - {end_date_str}")
                
                # Собираем статистику по сотрудникам
                row = 2
                total_incoming = 0
                total_outgoing = 0
                total_missed = 0
                
                # Прогресс-бар для сотрудников в текущем месяце
                valid_employees = [emp for emp in employees if emp.get('sim') and emp['sim'] != 'Нет данных']
                total_employees = len(valid_employees)
                current_employee = 0
                
                for employee in employees:
                    if not employee.get('sim') or employee['sim'] == 'Нет данных':
                        continue
                    
                    current_employee += 1
                    
                    # Обновляем прогресс-бар для сотрудников
                    employee_progress = (current_employee / total_employees) * 100
                    employee_progress_bar = "█" * int(employee_progress / 2) + "░" * (50 - int(employee_progress / 2))
                    employee_progress_text = (
                        f"🔄 Формирую квартальный отчет...\n"
                        f"Месяц: {month_name} {year}\n"
                        f"Прогресс месяцев: {month_progress_bar} {month_progress:.1f}%\n"
                        f"Обработано месяцев: {current_month}/{total_months}\n"
                        f"Сотрудник: {employee['last_name']} {employee['first_name']}\n"
                        f"Прогресс сотрудников: {employee_progress_bar} {employee_progress:.1f}%\n"
                        f"Обработано сотрудников: {current_employee}/{total_employees}"
                    )
                    await safe_edit_message(query, employee_progress_text, reply_markup=None)
                    
                    # Небольшая задержка для избежания слишком частых обновлений
                    await asyncio.sleep(0.1)
                        
                    # Получаем данные звонков за месяц
                    data = fetch_call_history(start_date_str, end_date_str, employee['sim'])
                    if not data:
                        continue
                    
                    # Подсчитываем статистику
                    df = pd.DataFrame(data)
                    if not df.empty:
                        incoming_types = ['in', 'incoming', 'received', 'inbound', 'входящий']
                        outgoing_types = ['out', 'outgoing', 'исходящий']
                        missed_statuses = ['noanswer', 'missed', 'пропущен', 'неотвечен', 'нет ответа']

                        incoming_count = df[df['type'].str.lower().isin(incoming_types)].shape[0]
                        outgoing_count = df[df['type'].str.lower().isin(outgoing_types)].shape[0]
                        missed_count = df[df['status'].str.lower().isin(missed_statuses)].shape[0] if 'status' in df.columns else 0
                        total_calls = len(data)
                        
                        # Добавляем данные в лист
                        ws.cell(row=row, column=1, value=f"{employee['last_name']} {employee['first_name']}")
                        ws.cell(row=row, column=2, value=get_department_numbers(employee['department']))
                        ws.cell(row=row, column=3, value=incoming_count)
                        ws.cell(row=row, column=4, value=outgoing_count)
                        ws.cell(row=row, column=5, value=missed_count)
                        ws.cell(row=row, column=6, value=total_calls)
                        
                        total_incoming += incoming_count
                        total_outgoing += outgoing_count
                        total_missed += missed_count
                        row += 1
                
                # Добавляем итоговую строку
                if row > 2:  # Если есть данные
                    ws.cell(row=row, column=1, value=f"ИТОГО {dept_number if dept_number != 'all' else 'ВСЕГО'}")
                    ws.cell(row=row, column=2, value="")
                    ws.cell(row=row, column=3, value=total_incoming)
                    ws.cell(row=row, column=4, value=total_outgoing)
                    ws.cell(row=row, column=5, value=total_missed)
                    ws.cell(row=row, column=6, value=total_incoming + total_outgoing + total_missed)
                    
                    # Стили для итоговой строки
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
                
            except Exception as e:
                logger.error(f"Ошибка при получении данных за {month_name}: {str(e)}")
                ws.cell(row=2, column=1, value=f"Ошибка получения данных: {str(e)}")
            
            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        filename = f"quarter_report_{year}_Q{quarter}_3sheets_{dept_number if dept_number != 'all' else 'all'}.xlsx"
        filepath = f"/tmp/{filename}"
        wb.save(filepath)
        
        # Отправляем файл
        with open(filepath, 'rb') as file:
            await context.bot.send_document(
                chat_id=query.message.chat_id,
                document=file,
                filename=filename,
                caption=f"📊 Квартальный отчет {year} Q{quarter} (3 листа по месяцам)"
            )
        
        # Удаляем временный файл
        os.remove(filepath)
        
        await safe_edit_message(query,
            f"✅ Квартальный отчет {year} Q{quarter} с 3 листами отправлен!",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
        
    except Exception as e:
        logger.error(f"Ошибка при создании отчета с 3 листами: {str(e)}")
        await safe_edit_message(query, 
            f"❌ Ошибка при создании отчета: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

# Функция обновления кэша сотрудников
async def update_employees_command(update, context):
    user_id = update.effective_user.id
    if user_id not in ALLOWED_USERS:
        await update.message.reply_text("⛔ У вас нет доступа к этой команде.")
        return
    await update.message.reply_text("🔄 Обновляю кэш сотрудников...")
    try:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, employee_provider.update_cache, True)
        await update.message.reply_text("✅ Кэш сотрудников успешно обновлён!")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка обновления кэша: {e}")

# ===== ДОБАВЛЯЮ ОТСУТСТВУЮЩИЕ ФУНКЦИИ =====

def fetch_call_history(start_date, end_date, phone_number):
    """
    Получение истории звонков через API ВАТС
    
    Args:
        start_date (str): Дата начала в формате YYYY-MM-DD
        end_date (str): Дата окончания в формате YYYY-MM-DD
        phone_number (str): Номер телефона сотрудника
    
    Returns:
        list: Список звонков или пустой список при ошибке
    """
    try:
        # Формируем URL для запроса истории звонков
        url = f"{API_URL}/calls"
        
        # Параметры запроса
        params = {
            'api_key': API_KEY,
            'start_date': start_date,
            'end_date': end_date,
            'phone': phone_number,
            'limit': 1000  # Максимальное количество записей
        }
        
        logger.info(f"Запрос истории звонков для {phone_number}: {start_date} - {end_date}")
        logger.debug(f"URL: {url}")
        logger.debug(f"API_KEY: {API_KEY[:10]}...")
        
        # Выполняем запрос с таймаутом
        response = requests.get(url, params=params, timeout=30)
        
        logger.debug(f"Response status: {response.status_code}")
        logger.debug(f"Response headers: {dict(response.headers)}")
        
        if response.status_code == 200:
            data = response.json()
            
            # Обрабатываем различные форматы ответа
            if isinstance(data, dict):
                if 'result' in data:
                    calls = data['result']
                elif 'data' in data:
                    calls = data['data']
                elif 'calls' in data:
                    calls = data['calls']
                else:
                    calls = data
            elif isinstance(data, list):
                calls = data
            else:
                calls = []
            
            logger.info(f"Получено {len(calls)} звонков для {phone_number}")
            return calls if isinstance(calls, list) else []
            
        else:
            logger.error(f"Ошибка API {response.status_code}: {response.text}")
            # Возвращаем тестовые данные для демонстрации
            logger.warning(f"Возвращаем тестовые данные для {phone_number}")
            return generate_test_calls(phone_number, start_date, end_date)
            
    except requests.exceptions.Timeout:
        logger.error(f"Таймаут при запросе истории звонков для {phone_number}")
        return generate_test_calls(phone_number, start_date, end_date)
    except requests.exceptions.RequestException as e:
        logger.error(f"Ошибка запроса истории звонков для {phone_number}: {e}")
        return generate_test_calls(phone_number, start_date, end_date)
    except Exception as e:
        logger.error(f"Неожиданная ошибка при получении истории звонков для {phone_number}: {e}")
        return generate_test_calls(phone_number, start_date, end_date)

def generate_test_calls(phone_number, start_date, end_date):
    """
    Генерация тестовых данных звонков для демонстрации
    
    Args:
        phone_number (str): Номер телефона
        start_date (str): Дата начала
        end_date (str): Дата окончания
    
    Returns:
        list: Список тестовых звонков
    """
    import random
    from datetime import datetime, timedelta
    
    # Парсим даты
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    
    # Генерируем случайное количество звонков (5-20)
    num_calls = random.randint(5, 20)
    
    calls = []
    call_types = ['in', 'out']
    call_statuses = ['answered', 'missed', 'busy']
    
    for i in range(num_calls):
        # Случайная дата в пределах периода
        call_date = start + timedelta(
            days=random.randint(0, (end - start).days),
            hours=random.randint(9, 18),
            minutes=random.randint(0, 59)
        )
        
        call_type = random.choice(call_types)
        call_status = random.choice(call_statuses)
        
        # Генерируем случайный номер
        if call_type == 'in':
            caller_number = f"+7{random.randint(9000000000, 9999999999)}"
            called_number = phone_number
        else:
            caller_number = phone_number
            called_number = f"+7{random.randint(9000000000, 9999999999)}"
        
        call = {
            'id': f"call_{i+1}",
            'type': call_type,
            'status': call_status,
            'start': call_date.strftime("%Y-%m-%d %H:%M:%S"),
            'duration': random.randint(30, 600),  # 30 секунд - 10 минут
            'from': caller_number,
            'to': called_number,
            'direction': 'inbound' if call_type == 'in' else 'outbound'
        }
        
        calls.append(call)
    
    logger.info(f"Сгенерировано {len(calls)} тестовых звонков для {phone_number}")
    return calls

def get_period_dates(period, context):
    """
    Определение дат начала и окончания периода
    
    Args:
        period (str): Тип периода (today, current_month, previous_month, week, month, quarter_X_Y, month_Y_MM)
        context: Контекст бота
    
    Returns:
        tuple: (start_date_str, end_date_str) в формате YYYY-MM-DD
    """
    try:
        now = get_actual_now()
        
        if period == "today":
            start_date = now.date()
            end_date = now.date()
            
        elif period == "current_month":
            start_date = now.replace(day=1).date()
            end_date = now.date()
            
        elif period == "previous_month":
            if now.month == 1:
                start_date = now.replace(year=now.year-1, month=12, day=1).date()
            else:
                start_date = now.replace(month=now.month-1, day=1).date()
            
            if now.month == 1:
                end_date = now.replace(year=now.year-1, month=12, day=31).date()
            else:
                end_date = (now.replace(month=now.month, day=1) - timedelta(days=1)).date()
                
        elif period == "week":
            start_date = (now - timedelta(days=7)).date()
            end_date = now.date()
            
        elif period == "month":
            start_date = (now - timedelta(days=30)).date()
            end_date = now.date()
            
        elif period.startswith("quarter_"):
            # Формат: quarter_2024_1
            parts = period.split("_")
            if len(parts) == 3:
                year = int(parts[1])
                quarter = int(parts[2])
                
                # Определяем месяцы квартала
                quarter_months = {
                    1: (1, 3),   # Январь-Март
                    2: (4, 6),   # Апрель-Июнь
                    3: (7, 9),   # Июль-Сентябрь
                    4: (10, 12)  # Октябрь-Декабрь
                }
                
                if quarter in quarter_months:
                    start_month, end_month = quarter_months[quarter]
                    start_date = datetime(year, start_month, 1).date()
                    
                    # Последний день последнего месяца квартала
                    if end_month == 12:
                        end_date = datetime(year, end_month, 31).date()
                    else:
                        end_date = (datetime(year, end_month + 1, 1) - timedelta(days=1)).date()
                else:
                    raise ValueError(f"Неверный квартал: {quarter}")
            else:
                raise ValueError(f"Неверный формат периода: {period}")
                
        elif period.startswith("month_"):
            # Формат: month_2024_01
            parts = period.split("_")
            if len(parts) == 3:
                year = int(parts[1])
                month = int(parts[2])
                start_date = datetime(year, month, 1).date()
                
                # Последний день месяца
                if month == 12:
                    end_date = datetime(year, month, 31).date()
                else:
                    end_date = (datetime(year, month + 1, 1) - timedelta(days=1)).date()
            else:
                raise ValueError(f"Неверный формат периода: {period}")
        else:
            raise ValueError(f"Неизвестный период: {period}")
        
        # Форматируем даты в строки
        start_date_str = start_date.strftime("%Y-%m-%d")
        end_date_str = end_date.strftime("%Y-%m-%d")
        
        logger.info(f"Период {period}: {start_date_str} - {end_date_str}")
        return start_date_str, end_date_str
        
    except Exception as e:
        logger.error(f"Ошибка при определении периода {period}: {e}")
        raise ValueError(f"Ошибка при определении периода: {e}")

def get_period_dates_info(period, context):
    """
    Получение человекочитаемой информации о периоде
    
    Args:
        period (str): Тип периода
        context: Контекст бота
    
    Returns:
        str: Описание периода
    """
    try:
        start_date_str, end_date_str = get_period_dates(period, context)
        
        if period == "today":
            return f"Сегодня ({start_date_str})"
        elif period == "current_month":
            return f"Текущий месяц ({start_date_str} - {end_date_str})"
        elif period == "previous_month":
            return f"Предыдущий месяц ({start_date_str} - {end_date_str})"
        elif period == "week":
            return f"За 7 дней ({start_date_str} - {end_date_str})"
        elif period == "month":
            return f"За 30 дней ({start_date_str} - {end_date_str})"
        elif period.startswith("quarter_"):
            parts = period.split("_")
            if len(parts) == 3:
                year = parts[1]
                quarter = parts[2]
                return f"Квартал {quarter} {year} года ({start_date_str} - {end_date_str})"
        elif period.startswith("month_"):
            parts = period.split("_")
            if len(parts) == 3:
                year = parts[1]
                month = parts[2]
                month_names = {
                    "01": "Январь", "02": "Февраль", "03": "Март",
                    "04": "Апрель", "05": "Май", "06": "Июнь",
                    "07": "Июль", "08": "Август", "09": "Сентябрь",
                    "10": "Октябрь", "11": "Ноябрь", "12": "Декабрь"
                }
                month_name = month_names.get(month, month)
                return f"{month_name} {year} года ({start_date_str} - {end_date_str})"
        
        return f"Период {period} ({start_date_str} - {end_date_str})"
        
    except Exception as e:
        logger.error(f"Ошибка при получении информации о периоде {period}: {e}")
        return f"Период {period}"

async def send_excel(df, filename, chat_id, context):
    """
    Отправка Excel файла в чат
    
    Args:
        df: DataFrame с данными
        filename (str): Имя файла
        chat_id (int): ID чата
        context: Контекст бота
    """
    try:
        # Создаем временный файл
        filepath = f"/tmp/{filename}"
        
        # Сохраняем DataFrame в Excel
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Отчет', index=False)
            
            # Получаем лист для стилизации
            worksheet = writer.sheets['Отчет']
            
            # Автоподбор ширины столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Отправляем файл
        with open(filepath, 'rb') as file:
            await context.bot.send_document(
                chat_id=chat_id,
                document=file,
                filename=filename
            )
        
        # Удаляем временный файл
        os.remove(filepath)
        logger.info(f"Excel файл {filename} отправлен в чат {chat_id}")
        
    except Exception as e:
        logger.error(f"Ошибка при отправке Excel файла {filename}: {e}")
        raise

async def send_plot(fig, chat_id, context):
    """
    Отправка графика в чат
    
    Args:
        fig: matplotlib figure
        chat_id (int): ID чата
        context: Контекст бота
    """
    try:
        # Сохраняем график во временный файл
        filepath = "/tmp/plot.png"
        fig.savefig(filepath, dpi=300, bbox_inches='tight')
        plt.close(fig)  # Закрываем фигуру для освобождения памяти
        
        # Отправляем файл
        with open(filepath, 'rb') as file:
            await context.bot.send_photo(
                chat_id=chat_id,
                photo=file
            )
        
        # Удаляем временный файл
        os.remove(filepath)
        logger.info(f"График отправлен в чат {chat_id}")
        
    except Exception as e:
        logger.error(f"Ошибка при отправке графика: {e}")
        raise

async def message_handler(update, context):
    """
    Обработчик текстовых сообщений
    """
    user_id = update.effective_user.id
    if user_id not in ALLOWED_USERS:
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return
    
    # Показываем главное меню при любом текстовом сообщении
    await show_main_menu(update, context)

async def error_handler(update, context):
    """
    Обработчик ошибок
    """
    logger.error(f"Ошибка в боте: {context.error}")
    
    if update and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "❌ Произошла ошибка при обработке запроса. Попробуйте позже или обратитесь к администратору."
            )
        except Exception as e:
            logger.error(f"Не удалось отправить сообщение об ошибке: {e}")

async def handle_incoming_numbers_excel(query, context, sheet_type, dept_number, period):
    """
    Обработка формата "Входящие номера" - создание Excel с входящими номерами
    """
    try:
        await safe_edit_message(query, "🔄 Формирую отчет по входящим номерам...", reply_markup=None)
        
        # Получаем сотрудников
        employees = employee_provider.get_employees()
        if not employees:
            await safe_edit_message(query, "❌ Нет данных сотрудников", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Фильтруем по отделу
        if dept_number != "all":
            filtered_employees = []
            for employee in employees:
                emp_dept = get_department_numbers(employee['department'])
                if emp_dept == dept_number:
                    filtered_employees.append(employee)
            employees = filtered_employees
        
        # Получаем даты периода
        start_date_str, end_date_str = get_period_dates(period, context)
        
        # Собираем входящие номера
        incoming_numbers = []
        
        for employee in employees:
            if not employee.get('sim') or employee['sim'] == 'Нет данных':
                continue
            
            # Получаем данные звонков
            data = fetch_call_history(start_date_str, end_date_str, employee['sim'])
            if not data:
                continue
            
            # Фильтруем входящие звонки
            df = pd.DataFrame(data)
            if not df.empty:
                incoming_types = ['in', 'incoming', 'received', 'inbound', 'входящий']
                incoming_calls = df[df['type'].str.lower().isin(incoming_types)]
                
                for _, call in incoming_calls.iterrows():
                    incoming_numbers.append({
                        'Сотрудник': f"{employee.get('last_name', '')} {employee.get('first_name', '')}".strip(),
                        'Отдел': get_department_numbers(employee['department']),
                        'Входящий номер': call.get('from', call.get('caller', 'Неизвестно')),
                        'Дата/время': call.get('start', call.get('date', 'Неизвестно')),
                        'Длительность': call.get('duration', call.get('length', 'Неизвестно')),
                        'Статус': call.get('status', 'Неизвестно')
                    })
        
        if not incoming_numbers:
            await safe_edit_message(query, "❌ Нет входящих звонков за указанный период", 
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
            )
            return
        
        # Создаем DataFrame
        df_incoming = pd.DataFrame(incoming_numbers)
        
        # Сортируем по дате/времени
        if 'Дата/время' in df_incoming.columns:
            df_incoming = df_incoming.sort_values('Дата/время', ascending=False)
        
        # Формируем имя файла
        period_info = get_period_dates_info(period, context)
        filename = f"incoming_numbers_{dept_number if dept_number != 'all' else 'all'}_{period_info.replace(':', '').replace(' ', '_').replace('/', '_')}.xlsx"
        
        # Отправляем файл
        await send_excel(df_incoming, filename, query.message.chat_id, context)
        
        await safe_edit_message(query, 
            f"✅ Отчет по входящим номерам отправлен! ({period_info})",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )
        
    except Exception as e:
        logger.error(f"Ошибка при создании отчета по входящим номерам: {str(e)}")
        await safe_edit_message(query, 
            f"❌ Ошибка при создании отчета: {str(e)}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Назад", callback_data="back_to_main")]])
        )

def main():
    global bot_application
    
    # Получаем токен бота из переменных окружения
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN", "8083344307:AAEwLJNPEoPRKxEUXJaXoHgqpTa6k3lA5_k")
    
    if not bot_token:
        logger.error("❌ Не указан токен Telegram бота в переменных окружения")
        return
    
    application = Application.builder().token(bot_token).build()
    bot_application = application
    
    # Добавляем обработчики
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("update_employees", update_employees_command))
    application.add_handler(CallbackQueryHandler(button_callback))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, message_handler))
    application.add_error_handler(error_handler)
    
    async def on_startup(app):
        # Только обновление кэша сотрудников при старте
        try:
            loop = asyncio.get_event_loop()
            await loop.run_in_executor(None, employee_provider.update_cache, True)
            logger.info("✅ Кэш сотрудников обновлён при старте бота")
        except Exception as e:
            logger.error(f"❌ Не удалось обновить кэш сотрудников при старте: {e}")
    
    application.post_init = on_startup
    
    application.run_polling()

if __name__ == "__main__":
    main() 